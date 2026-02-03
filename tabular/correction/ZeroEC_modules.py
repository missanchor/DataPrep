import os
import re
import json
import time
import ast
import random
import copy
import numpy as np
import pandas as pd
from typing import List
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

from langchain_community.vectorstores import FAISS
from langchain_community.vectorstores.utils import DistanceStrategy
from langchain_core.embeddings import Embeddings
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain.prompts import HumanMessagePromptTemplate, ChatPromptTemplate, SystemMessagePromptTemplate
from langchain_openai import ChatOpenAI
from fast_sentence_transformers import FastSentenceTransformer
from sklearn.cluster import MiniBatchKMeans
from sklearn.metrics import mutual_info_score
from joblib import Parallel, delayed
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# ==========================================
# Classes
# ==========================================

class Output(BaseModel):
    chain_of_thought_for_correction: str = Field(...,
                                                 description="The chain_of_thought that led to the proposed correction")
    correction: dict = Field(..., description="the most probable correction for the dirty value")


class myEmbeddings(Embeddings):
    def __init__(self, modelPath):
        self.model = FastSentenceTransformer(modelPath, device="cuda", quantize=True)

    def embed_documents(self, texts: List[str]) -> np.ndarray:
        return self.model.encode(texts, normalize_embeddings=True, batch_size=64, show_progress_bar=True)

    def embed_query(self, text: str) -> np.ndarray:
        return self.model.encode(text, normalize_embeddings=True)


# ==========================================
# 工具函数
# ==========================================

def get_folder_name(base_path):
    if not os.path.exists(base_path):
        os.makedirs(base_path)
    existing_folders = [d for d in os.listdir(base_path) if
                        os.path.isdir(os.path.join(base_path, d)) and d.startswith('run-')]
    max_run = 0
    for folder in existing_folders:
        try:
            run_number = int(folder.split('-')[1])
            if run_number > max_run: max_run = run_number
        except:
            pass
    next_run_path = os.path.join(base_path, f'run-{max_run + 1}')
    os.makedirs(next_run_path)
    return next_run_path


def load_prompts(*file_paths: str) -> tuple:
    return tuple(open(path, 'r', encoding='utf-8').read() for path in file_paths)


def form_examples(examps) -> str:
    few_shot_examps_str = ''
    for examp in examps:
        examp_str = 'human: ' + json.dumps(examp['input']) + '\n' + 'ai: ' + json.dumps(examp['output']) + '\n'
        few_shot_examps_str += examp_str
    return few_shot_examps_str


def load_examples(file_path: str) -> str:
    with open(file_path, 'r', encoding='utf-8') as file:
        content = ast.literal_eval(file.read())
    return form_examples(content)


def _clean_json_string(raw_str):
    if not raw_str: return None
    clean_str = re.sub(r'```json|```', '', raw_str).strip()
    return clean_str.replace('', '')


def format_row(row, header):
    return '{' + ', '.join(f'"{col}": "{val}"' for col, val in zip(header, row)) + '}'


def format_row_2(value, key, detection_row):
    result = {key[i]: value[i] for i in range(len(value)) if detection_row[i] == 0}
    return json.dumps(result)


def sort_dicts(dict_list, key1, key2, key3):
    return sorted(dict_list,
                  key=lambda x: (x.get(key1, float('inf')), x.get(key2, float('inf')), x.get(key3, float('inf'))))


def initialize_llm(model_name, api_base, api_key, temperature=0):
    return ChatOpenAI(model=model_name, base_url=api_base, api_key=api_key, temperature=temperature, max_tokens=4096,
                      model_kwargs={"response_format": {"type": "json_object"}})


def harmonic_mean(a, b):
    return (2 * a * b) / (a + b) if (a + b) > 0 else 0


# ==========================================
# 初始化函数
# ==========================================

def load_all_prompts(prompt_dir):
    def lp(name): return os.path.join(prompt_dir, name)

    prompts = {}
    prompts['general_examples_str'] = load_examples(lp('examples.txt'))
    prompts['examples_auto_cot_str'] = load_examples(lp('examples_for_AutoCoT_with_error_type.txt'))
    prompts['sys'], prompts['human'] = load_prompts(lp('SystemMessage-2.txt'), lp('HumanMessage.txt'))
    prompts['sys_auto_cot'], prompts['human_auto_cot_large'], prompts['human_auto_cot_small'] = load_prompts(
        lp('SystemMessage_for_AutoCoT_with_error_type.txt'), lp('HumanMessage_for_AutoCoT_large.txt'),
        lp('HumanMessage_for_AutoCoT_small.txt')
    )
    prompts['sys_code_generation'], prompts['human_code_generation'] = load_prompts(
        lp('SystemMessage_code_generation.txt'), lp('HumanMessage_code_generation.txt')
    )
    prompts['sys_fd_generation'], prompts['human_fd_generation'] = load_prompts(
        lp('SystemMessage_fd_generation.txt'), lp('HumanMessage_fd_generation.txt')
    )
    return prompts


def load_datasets(params):
    d = {}
    d['clean_data'] = pd.read_csv(params['clean_data_path'], dtype=str, encoding='utf-8').fillna('null')
    d['dirty_data'] = pd.read_csv(params['dirty_data_path'], dtype=str, encoding='utf-8').fillna('null')
    d['detection'] = pd.read_csv(params['detection_path'])
    d['row_count'], d['column_count'] = d['dirty_data'].shape
    d['header'] = d['dirty_data'].columns.tolist()
    return d


def init_all_models(params):
    m = {}
    m['llm'] = initialize_llm(params['MODEL_NAME'], params['OPENAI_API_BASE'], params['OPENAI_API_KEY'], 0.5)
    m['llm_auto_cot'] = initialize_llm(params['MODEL_NAME'], params['OPENAI_API_BASE'], params['OPENAI_API_KEY'], 0.5)
    m['llm_code_generation'] = initialize_llm(params['MODEL_NAME'], params['OPENAI_API_BASE'], params['OPENAI_API_KEY'],
                                              0)
    m['llm_fd_generation'] = initialize_llm(params['MODEL_NAME'], params['OPENAI_API_BASE'], params['OPENAI_API_KEY'],
                                            0)
    return m


# ==========================================
# 核心算法部件
# ==========================================

def calc_mi_2(df, target_column):
    def calculate_mi_optimized(column, target_column, df):
        if column == target_column:
            return mutual_info_score(df[target_column], df[column])
        else:
            counts = df.groupby([target_column, column]).size()
            filtered_counts = counts[counts > 1].reset_index()
            if filtered_counts.empty: return 0
            return mutual_info_score(filtered_counts[target_column], filtered_counts[column])

    mutual_info_list = Parallel(n_jobs=-1)(
        delayed(calculate_mi_optimized)(col, target_column, df) for col in df.columns
    )
    max_mutual_info = max(mutual_info_list) if mutual_info_list else 0
    normalized_mi = [mi / max_mutual_info if max_mutual_info != 0 else 0 for mi in mutual_info_list]
    return normalized_mi


def select_repair_candidates(embeddings_matrix: np.ndarray, detection: pd.DataFrame, num_clusters: int) -> list:
    mask_rows = detection.sum(axis=1) > 0
    filtered_embeddings = embeddings_matrix[mask_rows]
    filtered_detection = detection[mask_rows]
    original_indices = np.where(mask_rows)[0]

    m, n, l = filtered_embeddings.shape
    mask = filtered_detection.astype(bool)
    masked_embeddings = filtered_embeddings * mask.values[..., np.newaxis]
    reshaped_embeddings = masked_embeddings.reshape((m, n * l))

    kmeans = MiniBatchKMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(reshaped_embeddings)

    selected_indices = []
    covered_columns = set()

    for i in range(num_clusters):
        cluster_mask = clusters == i
        if np.any(cluster_mask):
            cluster_detection = filtered_detection.iloc[cluster_mask]
            cluster_original_indices = original_indices[cluster_mask]
            best_index = None
            best_new_coverage = -1
            best_orig_index = None

            for idx, orig_idx in zip(cluster_detection.index, cluster_original_indices):
                row = cluster_detection.loc[idx]
                new_columns = set(row[row == 1].index) - covered_columns
                if len(new_columns) > best_new_coverage:
                    best_new_coverage = len(new_columns)
                    best_index = idx
                    best_orig_index = orig_idx

            if best_index is not None:
                selected_indices.append(best_orig_index)
                covered_columns.update(
                    set(cluster_detection.loc[best_index][cluster_detection.loc[best_index] == 1].index))
    return selected_indices


def build_retriever_3(dirty_data, detection, vectors_matrix, target_column, dirty_data_human_repaired, embeddingModel):
    start_time = time.time()
    texts = dirty_data.values.tolist()
    print(f"Step 1: Converting dirty data to texts")

    step_start_time = time.time()
    normMI = calc_mi_2(dirty_data, target_column)
    normMI = np.round(normMI, decimals=1)
    normMI = np.array(normMI)
    threshold = 0.5

    indices = np.where(normMI >= threshold)[0]
    CoE = normMI[indices]
    embeddings_matrix_col_filtered = vectors_matrix[:, indices, :]
    texts_col_filtered = [[row[col] for col in indices] for row in texts]
    detection_col_filtered = detection.iloc[:, indices]

    formatted_rows = []
    header_col_filtered = detection_col_filtered.columns
    for text_row, detection_row in zip(texts_col_filtered, detection_col_filtered.values.tolist()):
        formatted_rows.append(format_row_2(text_row, header_col_filtered, detection_row))

    embeddings_matrix_col_row_filtered_reshaped = embeddings_matrix_col_filtered.reshape(
        embeddings_matrix_col_filtered.shape[0], -1)

    paired_data = list(zip(formatted_rows, embeddings_matrix_col_row_filtered_reshaped))
    ids = [str(i) for i in range(len(dirty_data_human_repaired))]
    meta_data = [{'index': i} for i in range(len(dirty_data_human_repaired))]

    db = FAISS.from_embeddings(text_embeddings=paired_data,
                               embedding=embeddingModel,
                               metadatas=meta_data,
                               ids=ids,
                               distance_strategy=DistanceStrategy.DOT_PRODUCT)
    print(f"Total Time Taken for Retriever Build: {time.time() - start_time:.4f}s")
    return db, indices, CoE


def get_auto_cot(repair_list, column, retriever, CoE, indices, detection_filtered, params):
    rep_error_info = params['rep_error_info']
    sp_examps = params['sp_examps']
    dirty_data = params['dirty_data']
    clean_data = params['clean_data']
    embeddings_matrix = params['embeddings_matrix']
    detection_human_repaired = params['detection_human_repaired']
    indices_dict = params['indices_dict']
    header = params['header']
    llm_auto_cot = params['llm_auto_cot']

    rep_error_info[column] = {}
    human_message_small_template = HumanMessagePromptTemplate.from_template(params['human_auto_cot_small'])
    human_input = '['
    relevant_rows_list = []
    dirty_tuples_list = []
    dirty_values = []
    clean_values = []
    filtered_header = [header[i] for i in indices]

    column_detection = detection_human_repaired[column].values
    column_indices = indices_dict[column]
    column_sums = detection_human_repaired.iloc[:, column_indices].values.sum(axis=1)

    for row_idx in repair_list:
        if dirty_data.loc[row_idx, column] != clean_data.loc[row_idx, column]:
            dirty_value = dirty_data.loc[row_idx, column]
            dirty_values.append(dirty_value)
            clean_value = clean_data.loc[row_idx, column]
            clean_values.append(clean_value)

            dirty_row_data = dirty_data.iloc[:, indices].loc[row_idx]
            dirty_tuples_list.append(dirty_row_data.tolist())

            embeddings_row = embeddings_matrix[row_idx]
            embeddings_row_filtered = embeddings_row[indices].copy()
            for i in range(len(embeddings_row_filtered)):
                if detection_filtered.iloc[row_idx, i] == 1:
                    embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))

            embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
            embeddings_row_united = embeddings_row_filtered.flatten()
            relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united, k=30)

            relevant_rows_dict_list = [
                {
                    'page_content': row[0].page_content,
                    'index': idx,
                    'score': round(row[1], 2),
                    'target_column': column_detection[idx],
                    'sum': column_sums[idx]
                }
                for row in relevant_rows
                for idx in [row[0].metadata['index']]
            ]

            sorted_relevant_rows = sort_dicts(relevant_rows_dict_list, 'score', 'target_column', 'sum')
            relevant_clean_tuples = ''
            for row in sorted_relevant_rows[:3]:
                relevant_clean_tuples += row['page_content'] + '\n'
            relevant_rows_list.append(relevant_clean_tuples)

            human_input += '{' + human_message_small_template.format(
                Dirty_Tuple=format_row(dirty_row_data.tolist(), filtered_header),
                Erroneous_value='{' + f'"{column}": "{dirty_value}"' + '}',
                Relevant_clean_tuples=relevant_clean_tuples,
                Correction=clean_value
            ).content + '},'
    human_input += ']'

    prompt_auto_cot = ChatPromptTemplate(messages=[
        SystemMessagePromptTemplate.from_template(params['sys_auto_cot']),
        HumanMessagePromptTemplate.from_template(params['human_auto_cot_large']),
    ], partial_variables={"examples": params['examples_auto_cot_str']})

    chain_auto_cot = (prompt_auto_cot | llm_auto_cot)

    repair_result_list = []
    while True:
        try:
            response = chain_auto_cot.invoke({"human_input": human_input})
            repair_result = response.content
            repair_result = re.sub(r'```json|```', '', repair_result).strip().replace('\ufffd', '')
            try:
                temp_result = json.loads(repair_result)
            except:
                temp_result = ast.literal_eval(repair_result)

            if isinstance(temp_result, dict):
                repair_result_list = [temp_result]
            elif isinstance(temp_result, list):
                repair_result_list = temp_result
            elif isinstance(temp_result, str):
                second_parse = ast.literal_eval(temp_result)
                repair_result_list = [second_parse] if isinstance(second_parse, dict) else second_parse
            else:
                raise ValueError(f"Unknown return type: {type(temp_result)}")

            if isinstance(repair_result_list, list):
                break
            else:
                raise ValueError("Result is not a list")
        except Exception as e:
            print(f'Column {column} - Auto-CoT parsing failed: {e}')
            time.sleep(2)

    specific_examples_llm = []
    for idx, result in enumerate(repair_result_list):
        if isinstance(result, str):
            try:
                result = json.loads(result)
            except:
                try:
                    result = ast.literal_eval(result)
                except:
                    continue
        if not isinstance(result, dict): continue

        cot = result.get('chain_of_thought_for_correction', 'No reasoning provided.')
        err_type = result.get('error_type', 'General Error')
        dirty_tuple = dirty_tuples_list[idx]

        json_template = {
            "input": {
                "Dirty Data Tuple": format_row(dirty_tuple, filtered_header),
                "Erroneous Value": '{' + f'"{column}": "{dirty_values[idx]}"' + '}',
                "Relevant Clean Tuples": relevant_rows_list[idx]
            },
            "output": {
                "chain_of_thought_for_correction": cot,
                "correction": '{"' + column + '\": ' + '"' + clean_values[idx] + '"}'
            }
        }
        row_name = dirty_data.iloc[:, indices].loc[repair_list[idx]].name
        rep_error_info[column][row_name] = {
            'dirty_tuple': dirty_data.iloc[repair_list[idx]].to_dict(),
            # Note: using original dict might differ slightly from filtered
            'dirty_value': dirty_values[idx],
            'ground_truth': clean_values[idx],
            'error_analysis': cot,
            'error_type': err_type
        }
        specific_examples_llm.append(json_template)
    sp_examps[column] = form_examples(specific_examples_llm)


def sel_clean(num_clusters: int, detection, embeddings_matrix, dirty_data):
    mask_rows = detection.sum(axis=1) == 0
    filtered_embeddings = embeddings_matrix[mask_rows]
    m, n, l = filtered_embeddings.shape
    reshaped_embeddings = filtered_embeddings.reshape(m, n * l)
    if m == 0: return pd.DataFrame()
    if m < num_clusters: num_clusters = m
    kmeans = MiniBatchKMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(reshaped_embeddings)
    original_indices = np.where(mask_rows)[0]
    selected_indices = [
        original_indices[np.random.choice(np.where(clusters == cluster_id)[0])]
        for cluster_id in range(num_clusters)
        if np.any(clusters == cluster_id)
    ]
    return dirty_data.iloc[selected_indices]


def train_val_split(data: dict):
    keys = list(data.keys())
    random.shuffle(keys)
    split_index = int(len(data) * 0.5)
    train_keys = keys[:split_index]
    val_keys = keys[split_index:]
    return {key: data[key] for key in train_keys}, {key: data[key] for key in val_keys}


def clean_data_integration(clean_data: pd.DataFrame, rep_data_info: dict, dirty_data, detection):
    for column in dirty_data.columns:
        if detection[column].sum() > 0:
            for idx in range(len(clean_data)):
                row_name = clean_data.iloc[idx].name
                rep_data_info[column][row_name] = {}
                rep_data_info[column][row_name]['dirty_tuple'] = clean_data.iloc[idx].to_dict()
                rep_data_info[column][row_name]['dirty_value'] = clean_data.iloc[idx][column]
                rep_data_info[column][row_name][
                    'error_analysis'] = 'This is a clean value that does not need correction.'
                rep_data_info[column][row_name]['ground_truth'] = clean_data.iloc[idx][column]
                rep_data_info[column][row_name]['error_type'] = 'clean'


def code_generation(train_data: dict, column: str, codes: dict, params):
    examples = ''
    for error in train_data.values():
        dirty_value_str = error['dirty_value']
        analysis = error['error_analysis']
        clean_value_str = error['ground_truth']
        examples += f"[Erroneous_value: {dirty_value_str}\nError_analysis: {analysis}\nCorrect_value: {clean_value_str}]\n"

    prompt = ChatPromptTemplate(messages=[
        SystemMessagePromptTemplate.from_template(params['sys_code_generation']),
        HumanMessagePromptTemplate.from_template(params['human_code_generation']),
    ])
    chain = (prompt | params['llm_code_generation'])
    while True:
        try:
            raw_code = chain.invoke({'examples': examples}).content
            extracted_code = re.findall(r'```python(.*?)```', raw_code, re.DOTALL)
            if extracted_code:
                codes[column] = extracted_code[0]
            break
        except Exception as e:
            print('code generation failed', e)


def FD_generation(train_data: dict, column: str, FDs: dict, params):
    examples = ''
    for error in train_data.values():
        dirty_tuple_str = json.dumps(error['dirty_tuple'])
        dirty_value_str = f"{column}: {error['dirty_value']}"
        analysis = error['error_analysis']
        correct_value = f"{column}: {error['ground_truth']}"
        examples += f"Dirty_tuple: {dirty_tuple_str}\nErroneous_value: {dirty_value_str}, Error_analysis: {analysis}\nCorrect_value: {correct_value}]\n"

    prompt = ChatPromptTemplate(messages=[SystemMessagePromptTemplate.from_template(params['sys_fd_generation'])])
    chain = (prompt | params['llm_fd_generation'])
    while True:
        try:
            fd_raw = chain.invoke({'examples': examples}).content
            fd_raw = re.sub(r'```json|```python|```', '', fd_raw).strip()
            fd_raw = "\n".join([line.strip() for line in fd_raw.splitlines()])
            try:
                fd_dict = json.loads(fd_raw)
            except:
                fd_dict = ast.literal_eval(fd_raw)
            if isinstance(fd_dict, dict) and 'functional_dependency' in fd_dict:
                break
            else:
                raise ValueError("FD 格式不正确")
        except Exception as e:
            print(f'Column {column} - fd generation failed: {e}')
            time.sleep(1)
    FDs[column] = fd_dict


def correct(val):
    return val


def code_evaluation_execution(code, val_data, column, detection_human_repaired, corrections):
    try:
        local_vars = {}
        exec(code, globals(), local_vars)
        if 'correct' not in local_vars: return False
        correct_func = local_vars['correct']
    except Exception as e:
        print(f"Code compilation failed: {e}")
        return False

    flag = True
    for error in val_data.values():
        try:
            res = str(correct_func(error['dirty_value']))
            if res != error['ground_truth']:
                # print(f"expect {error['ground_truth']}, get {res}")
                flag = False
        except:
            flag = False

    if flag == False:
        print(f'{column} code failed')
        return flag

    mask = detection_human_repaired[column] == 1
    corrections.loc[mask, column] = corrections.loc[mask, column].apply(lambda x: str(correct_func(x)))
    corrections[column] = corrections[column].astype(str)
    print(f'{column} code passed')
    return True


def fd_evaluation_execution(fd: dict, val_data: dict, column: str, detection_human_repaired, dirty_data_human_repaired,
                            corrections):
    match = re.search(r"^(.+?)\s*→\s*(.+)$", fd['functional_dependency'])
    if match:
        attr1, attr2 = match.groups()
    else:
        return False

    if attr1 == 'None': return False
    flag = True
    filter_conditions = (detection_human_repaired[attr1] == 0) & (detection_human_repaired[attr2] == 0)

    for idx, error in val_data.items():
        if error['error_type'] == 'clean':
            pass

        if attr1 in error['dirty_tuple']:
            valid_rows = dirty_data_human_repaired[attr1] == error['dirty_tuple'][attr1]
        else:
            valid_rows = pd.Series([False] * len(dirty_data_human_repaired))

        attr2_values = dirty_data_human_repaired.loc[valid_rows & filter_conditions, attr2]
        if not attr2_values.empty:
            corrected_value = attr2_values.mode()[0]
        else:
            corrected_value = error['dirty_value']

        if corrected_value != error['ground_truth']:
            flag = False
            break

    if flag == False:
        print(f'{column} failed')
        return flag

    for idx in detection_human_repaired.index[detection_human_repaired[column] == 1]:
        val1 = dirty_data_human_repaired.at[idx, attr1]
        valid_rows = (dirty_data_human_repaired[attr1] == val1) & (detection_human_repaired[attr1] == 0) & (
                    detection_human_repaired[attr2] == 0)
        attr2_values = dirty_data_human_repaired.loc[valid_rows, attr2]
        if not attr2_values.empty:
            corrected_value = attr2_values.mode()[0]
        else:
            corrected_value = corrections.at[idx, column]
        corrections.at[idx, column] = corrected_value
    print(f'{column} fd passed')
    return True


def update_retriever(column, repair_list, retriever_dict, indices_dict, embeddings_matrix_only_repaired,
                     dirty_data_only_repaired, header):
    ids = [str(i) for i in repair_list]
    retriever_dict[column].delete(ids)
    embeddings_matrix_only_repaired_col_filtered = embeddings_matrix_only_repaired[:, indices_dict[column], :]
    texts = dirty_data_only_repaired.values.tolist()
    texts_col_filtered = [[row[col] for col in indices_dict[column]] for row in texts]
    header_col_filtered = [header[i] for i in indices_dict[column]]
    formatted_rows = []
    for text_row in texts_col_filtered:
        formatted_rows.append(format_row(text_row, header_col_filtered))
    meta_data = [{'index': i} for i in repair_list]
    paired_data = [(text, vector.flatten().tolist()) for text, vector in
                   zip(formatted_rows, embeddings_matrix_only_repaired_col_filtered)]
    retriever_dict[column].add_embeddings(text_embeddings=paired_data, metadatas=meta_data, ids=ids)


def repair_value(dirty_tuple, column, dirty_value, index_row, index_col, chain, params):
    indices_dict = params['indices_dict']
    header = params['header']
    retrieved_tuples = params['retrieved_tuples']
    corrections = params['corrections']
    logs = params['logs']

    filtered_tuple = dirty_tuple.iloc[indices_dict[column]]
    filtered_header = [header[i] for i in indices_dict[column]]
    dirty_tuple_filtered_str = format_row(filtered_tuple, filtered_header)
    dirty_value_str = '{' + f'"{column}": "{dirty_value}"' + '}'
    dirty_tuple_json = dirty_tuple.to_dict()
    correction = dirty_value
    result_json = dirty_value_str

    relevant_clean_tuples = retrieved_tuples[column][index_row]
    try_num = 0
    while True:
        try:
            raw_response = chain.invoke({
                'Dirty_Tuple': dirty_tuple_filtered_str,
                'Erroneous_value': dirty_value_str,
                'Relevant_clean_tuples': relevant_clean_tuples,
            })
            if raw_response is None: raise ValueError("LLM 返回为空 (None)")

            if isinstance(raw_response, str):
                cleaned_content = _clean_json_string(raw_response)
                try:
                    result_json = json.loads(cleaned_content)
                except:
                    match = re.search(r'"correction":\s*(\{[^}]+\})', cleaned_content)
                    if match:
                        result_json = {"correction": json.loads(match.group(1))}
                    else:
                        raise ValueError("无法解析 JSON 结构")
            else:
                result_json = raw_response

            if 'correction' not in result_json: raise KeyError("JSON 中缺少 'correction' 字段")
            correction = result_json['correction'].get(column, "null")
            break
        except Exception as e:
            tqdm.write(f"第 {index_row} 行 [{column}] 修复失败: {e}")
            try_num += 1
            if try_num >= 3:
                correction = dirty_value
                break
            time.sleep(2)

    corrections.iloc[index_row, index_col] = str(correction)
    log = {'Index': dirty_tuple_json.get('index', index_row),
           'Dirty_tuple': format_row(dirty_tuple, header),
           'Dirty_value': dirty_value_str,
           'Relevant_clean_tuples': relevant_clean_tuples,
           'Correction': str(result_json)
           }
    logs.append(log)


def cmp_mark(df_A, df_B, dirty_data_name, output_path, MODEL_NAME):
    df_A.fillna('null', inplace=True)
    df_B.fillna('null', inplace=True)
    difference = df_A.ne(df_B)
    diff_count = difference.sum().sum()
    diff_percent = diff_count / difference.size * 100

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df_B, index=False, header=True):
        ws.append(r)

    for col in range(difference.shape[1]):
        for row in range(difference.shape[0]):
            if difference.iloc[row, col]:
                cell = ws.cell(row=row + 2, column=col + 1)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    print(f"不一致元素个数: {diff_count}, 不一致元素所占百分比: {diff_percent:.2f}%")
    filename_without_extension = os.path.splitext(dirty_data_name)[0]
    save_path = os.path.join(output_path, f'{filename_without_extension}-corrected-marked_{MODEL_NAME}.xlsx')
    wb.save(save_path)


def calc_p_r_f(clean_data, dirty_data, corrected_data, output_path):
    mask_bc = corrected_data != dirty_data
    mask_ac = clean_data == corrected_data
    corrected_num = mask_bc.sum().sum()
    final_mask1 = mask_ac & mask_bc
    right_corrected_num1 = final_mask1.sum().sum()
    Precision = right_corrected_num1 / corrected_num

    mask_ab = clean_data != dirty_data
    dirty_num = mask_ab.sum().sum()
    final_mask2 = mask_ab & mask_ac
    right_corrected_num2 = final_mask2.sum().sum()
    Recall = right_corrected_num2 / dirty_num
    F1 = harmonic_mean(Precision, Recall)

    with open(os.path.join(output_path, 'output.txt'), 'a', encoding='utf-8') as f_output:
        print(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}')
        f_output.write(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}\n')


# ==========================================
# 流程包装函数
# ==========================================

def run_embedding_and_selection(params):
    print('embedding...')
    elements_list = params['dirty_data'].values.flatten().tolist()
    params['embeddingModel'] = myEmbeddings(params['EMBEDDING_MODEL_PATH'])
    embeddings = params['embeddingModel'].embed_documents(elements_list)
    embedding_dimension = len(embeddings[0])
    params['embeddings_matrix'] = embeddings.reshape(params['row_count'], params['column_count'], embedding_dimension)
    print('embedding done')

    print('selecting repair candidates...')
    params['repair_list'] = select_repair_candidates(params['embeddings_matrix'], params['detection'],
                                                     params['human_repair_num'])
    print(params['repair_list'])


def simulate_human_repair(params):
    human_repair_num = params['human_repair_num']
    repair_list = params['repair_list']
    dirty_data = params['dirty_data']
    clean_data = params['clean_data']
    detection = params['detection']
    embeddingModel = params['embeddingModel']
    column_count = params['column_count']
    embedding_dimension = params['embeddings_matrix'].shape[2]

    dirty_data_human_repaired = dirty_data.copy()
    dirty_data_human_repaired.iloc[repair_list] = clean_data.iloc[repair_list]
    params['dirty_data_human_repaired'] = dirty_data_human_repaired

    detection_human_repaired = detection.copy()
    detection_human_repaired.iloc[repair_list] = 0
    params['detection_human_repaired'] = detection_human_repaired

    params['corrections'].iloc[repair_list] = clean_data.iloc[repair_list]

    dirty_data_only_repaired = clean_data.iloc[repair_list]
    params['dirty_data_only_repaired'] = dirty_data_only_repaired

    elements_list_only_repaired = dirty_data_only_repaired.values.flatten().tolist()
    embeddings_only_repaired = embeddingModel.embed_documents(elements_list_only_repaired)
    params['embeddings_matrix_only_repaired'] = embeddings_only_repaired.reshape(human_repair_num, column_count,
                                                                                 embedding_dimension)


def run_auto_cot_generation(params):
    detection = params['detection']
    with ThreadPoolExecutor(max_workers=params['max_workers']) as executor:
        futures = []
        for col_idx, column in enumerate(detection.columns):
            if detection[column].sum() > 0:
                retriever, indices, CoE = build_retriever_3(
                    params['dirty_data'], detection, params['embeddings_matrix'], column,
                    params['dirty_data_human_repaired'], params['embeddingModel']
                )
                params['retriever_dict'][column] = retriever
                params['indices_dict'][column] = indices
                params['CoE_dict'][column] = CoE
                detection_filtered = detection.iloc[:, indices]

                future = executor.submit(get_auto_cot,
                                         params['repair_list'], column, retriever, CoE, indices, detection_filtered,
                                         params)
                futures.append(future)
        for future in tqdm(as_completed(futures), total=len(futures), desc="Auto-CoT 生成进度"):
            try:
                future.result()
            except Exception as e:
                print(f"Auto-CoT 任务出错: {e}")


def run_code_fd_generation(params):
    # Select clean data and integrate
    rep_clean_data = sel_clean(params['human_repair_num'], params['detection'], params['embeddings_matrix'],
                               params['dirty_data'])
    import copy
    rep_data_info = copy.deepcopy(params['rep_error_info'])
    if not rep_clean_data.empty:
        clean_data_integration(rep_clean_data, rep_data_info, params['dirty_data'], params['detection'])

    with ThreadPoolExecutor(max_workers=1) as executor:
        futures = []
        for column in params['detection'].columns:
            if params['detection'][column].sum() > 0:
                formatting_issue = True
                for error in params['rep_error_info'][column].values():
                    if error['error_type'] not in ['clean', 'Formatting Issue']:
                        formatting_issue = False
                        break
                params['train_data'][column], params['val_data'][column] = train_val_split(rep_data_info[column])

                if formatting_issue:
                    future = executor.submit(code_generation, params['train_data'][column], column, params['codes'],
                                             params)
                else:
                    future = executor.submit(FD_generation, params['train_data'][column], column, params['fds'], params)
                futures.append(future)
        for future in as_completed(futures): future.result()


def run_code_fd_execution(params):
    for column in params['codes'].keys():
        code_evaluation_execution(params['codes'][column], params['val_data'][column], column,
                                  params['detection_human_repaired'], params['corrections'])
    for column in params['fds'].keys():
        fd_evaluation_execution(params['fds'][column], params['val_data'][column], column,
                                params['detection_human_repaired'], params['dirty_data_human_repaired'],
                                params['corrections'])


def run_retriever_update(params):
    detection_human_repaired = params['detection_human_repaired']
    for column in detection_human_repaired.columns:
        if detection_human_repaired[column].sum() > 0:
            update_retriever(column, params['repair_list'], params['retriever_dict'], params['indices_dict'],
                             params['embeddings_matrix_only_repaired'], params['dirty_data_only_repaired'],
                             params['header'])


def run_retrieval(params):
    detection_human_repaired = params['detection_human_repaired']
    # Create a copy to track what still needs LLM repair (items not fixed by code/FD)
    detection_human_repaired_copy = detection_human_repaired.copy()
    mask = params['dirty_data_human_repaired'] != params['corrections']
    detection_human_repaired_copy[mask] = 0
    params['detection_human_repaired_copy'] = detection_human_repaired_copy

    for column in params['dirty_data_human_repaired'].columns:
        if detection_human_repaired[column].sum() > 0:
            retriever = params['retriever_dict'][column]
            indices = params['indices_dict'][column]
            CoE = params['CoE_dict'][column]
            temp = detection_human_repaired.iloc[:, indices]

            for row_idx in range(len(detection_human_repaired)):
                if detection_human_repaired_copy.at[row_idx, column] == 1:
                    column_detection = detection_human_repaired[column].values
                    column_indices = params['indices_dict'][column]
                    column_sums = detection_human_repaired.iloc[:, column_indices].values.sum(axis=1)

                    relevant_clean_tuples = ''
                    embeddings_row = params['embeddings_matrix'][row_idx]
                    embeddings_row_filtered = embeddings_row[indices].copy()

                    mask_emb = temp.iloc[row_idx].astype(bool).values
                    if mask_emb.any(): embeddings_row_filtered[mask_emb] = 0

                    embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
                    embeddings_row_united = embeddings_row_filtered.flatten()

                    relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united,
                                                                                     k=30)

                    relevant_rows_dict_list = [
                        {
                            'page_content': row[0].page_content,
                            'index': idx,
                            'score': round(row[1], 2),
                            'target_column': column_detection[idx],
                            'sum': column_sums[idx]
                        }
                        for row in relevant_rows for idx in [row[0].metadata['index']]
                    ]

                    sorted_relevant_rows_dict_list = sort_dicts(relevant_rows_dict_list, 'score', 'target_column',
                                                                'sum')
                    for row in sorted_relevant_rows_dict_list[:3]:
                        relevant_clean_tuples += row['page_content'] + '\n'

                    if column not in params['retrieved_tuples']: params['retrieved_tuples'][column] = {}
                    params['retrieved_tuples'][column][row_idx] = relevant_clean_tuples


def run_llm_repair(params):
    detection_human_repaired = params['detection_human_repaired']
    detection_human_repaired_copy = params['detection_human_repaired_copy']

    with ThreadPoolExecutor(max_workers=params['max_workers']) as executor:
        futures = []
        for col_idx, column in enumerate(detection_human_repaired.columns):
            if detection_human_repaired_copy[column].sum() > 0:
                prompt = ChatPromptTemplate(
                    messages=[
                        SystemMessagePromptTemplate.from_template(params['sys']),
                        HumanMessagePromptTemplate.from_template(params['human']),
                    ],
                    partial_variables={
                        "general_examples": params['general_examples_str'],
                        "specific_examples": params['sp_examps'][column],
                        "format_instructions": params['parser'].get_format_instructions()
                    }
                )
                params['prompt_dict'][column] = prompt
                chain = (prompt | params['llm'] | params['parser'])

                for row_idx in range(len(detection_human_repaired)):
                    if detection_human_repaired_copy.at[row_idx, column] == 1:
                        dirty_tuple = params['dirty_data_human_repaired'].iloc[row_idx]
                        dirty_value = params['dirty_data_human_repaired'].at[row_idx, column]
                        future = executor.submit(repair_value, dirty_tuple, column, dirty_value, row_idx, col_idx,
                                                 chain, params)
                        futures.append(future)
        for future in tqdm(as_completed(futures), total=len(futures), desc="修复进度"):
            try:
                future.result()
            except Exception as e:
                print(f"修复任务出错: {e}")


def run_evaluation(params, execution_time):
    logs = sorted(params['logs'], key=lambda x: int(x['Index']))
    params['corrections'].to_csv(os.path.join(params['output_path'], 'corrections.csv'), encoding='utf-8',
                                 index=False)
    cmp_mark(params['clean_data'], params['corrections'], os.path.basename(params['dirty_data_path']),
             params['output_path'], params['MODEL_NAME'])
    calc_p_r_f(params['clean_data'], params['dirty_data_human_repaired'], params['corrections'],
               params['output_path'])

    with open(os.path.join(params['output_path'], 'output.txt'), 'a', encoding='utf-8') as f_output:
        f_output.write(f"execution time: {execution_time} seconds\n")